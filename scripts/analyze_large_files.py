# -*- coding: utf-8 -*-
"""
Large File Analysis Script
Analyze structure of large CSV/XLSX files without loading everything into memory
"""

import csv
from pathlib import Path
import sys

def analyze_csv(file_path, sample_rows=10):
    """Analyze CSV file structure"""
    file_path = Path(file_path)

    if not file_path.exists():
        print(f"[SKIP] File not found: {file_path.name}")
        return None

    print(f"\n{'='*70}")
    print(f"FILE: {file_path.name}")
    print(f"SIZE: {file_path.stat().st_size / (1024*1024):.2f} MB")
    print(f"{'='*70}")

    try:
        # Try different encodings
        for encoding in ['utf-8-sig', 'cp949', 'utf-8']:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    reader = csv.reader(f)

                    # Get header
                    header = next(reader)

                    # Count rows and get samples
                    row_count = 0
                    samples = []

                    for i, row in enumerate(reader):
                        row_count += 1
                        if i < sample_rows:
                            samples.append(row)

                        # Progress indicator for large files
                        if row_count % 100000 == 0:
                            print(f"  Reading... {row_count:,} rows")

                print(f"\n[SUCCESS] Encoding: {encoding}")
                print(f"\nStructure:")
                print(f"  Rows: {row_count:,}")
                print(f"  Columns: {len(header)}")

                print(f"\nColumns ({len(header)}):")
                for i, col in enumerate(header, 1):
                    print(f"  {i}. {col}")

                print(f"\nFirst {min(sample_rows, len(samples))} Data Rows:")
                for i, row in enumerate(samples, 1):
                    print(f"  [{i}] {', '.join(str(x) for x in row[:5])}...")

                return {
                    'encoding': encoding,
                    'rows': row_count,
                    'columns': len(header),
                    'header': header,
                    'samples': samples
                }

            except UnicodeDecodeError:
                continue
            except Exception as e:
                print(f"[ERROR] {encoding}: {str(e)[:50]}")
                continue

        print("[FAIL] All encodings failed")
        return None

    except Exception as e:
        print(f"[ERROR] {str(e)}")
        return None

def analyze_xlsx(file_path):
    """Analyze XLSX file structure (requires openpyxl)"""
    try:
        import openpyxl
        file_path = Path(file_path)

        if not file_path.exists():
            print(f"[SKIP] File not found: {file_path.name}")
            return None

        print(f"\n{'='*70}")
        print(f"FILE: {file_path.name}")
        print(f"SIZE: {file_path.stat().st_size / (1024*1024):.2f} MB")
        print(f"{'='*70}")

        # Load workbook
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        print(f"\nWorkbook Info:")
        print(f"  Sheets: {len(wb.sheetnames)}")
        print(f"  Sheet Names: {', '.join(wb.sheetnames)}")

        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")

            # Get dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            print(f"  Dimensions: {max_row:,} rows x {max_col} columns")

            # Get header
            header = [cell.value for cell in sheet[1]]
            print(f"  Columns ({max_col}):")
            for i, col in enumerate(header, 1):
                if i <= 10:  # First 10 columns only
                    print(f"    {i}. {col}")
            if max_col > 10:
                print(f"    ... and {max_col - 10} more columns")

            # Sample rows
            print(f"\n  First 5 Data Rows:")
            for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=6, values_only=True), 1):
                row_preview = ', '.join(str(x) for x in list(row)[:5])
                print(f"    [{i}] {row_preview}...")

        wb.close()
        return {'sheets': wb.sheetnames, 'analyzed': True}

    except ImportError:
        print("[ERROR] openpyxl not installed. Install with: pip install openpyxl")
        return None
    except Exception as e:
        print(f"[ERROR] {str(e)}")
        return None

# Files to analyze
large_csvs = [
    r"c:\PROGRAMMING\2026_1_SPRING\metropy\inputs\elderly_hourly_station_daily_20231231.csv",
    r"c:\PROGRAMMING\2026_1_SPRING\metropy\inputs\hourly_line_station_cnt.csv",
]

xlsx_files = [
    r"c:\PROGRAMMING\2026_1_SPRING\metropy\inputs\congestion_line9_weekday_weekend_station_hour.xlsx",
    r"c:\PROGRAMMING\2026_1_SPRING\metropy\inputs\line_station_ID.xlsx",
]

print("\n" + "="*70)
print("ANALYZING LARGE CSV FILES")
print("="*70)

for file_path in large_csvs:
    result = analyze_csv(file_path, sample_rows=5)

print("\n" + "="*70)
print("ANALYZING XLSX FILES")
print("="*70)

for file_path in xlsx_files:
    result = analyze_xlsx(file_path)

print("\n\n[DONE] Analysis complete!")
